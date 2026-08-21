"""A tiny change-history export, shaped like Amazon's.

The golden tests run against a real export that cannot be committed. This one
is small and synthetic: enough for the hosted app's tests to upload something
and get a real analysis back, without depending on customer data.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

HEADER = ["Change level type", "Change level name", "Campaign", "Change type",
          "From", "To", "From (numeric)", "To (numeric)", "Date and time (ISO)"]

DATE = "2026-08-06"

ROWS = [
    # Out of budget mid-morning, never recovers.
    ("Campaign", "Chronic Offender", "Chronic Offender", "Campaign created", "", "", None, None, f"{DATE}T00:00:00"),
    ("Campaign", "Chronic Offender", "Chronic Offender", "Campaign daily budget", "$50.00", "$50.00", 50, 50, f"{DATE}T00:01:00"),
    ("Campaign", "Chronic Offender", "Chronic Offender", "Campaign status", "Out of budget", "In budget", None, None, f"{DATE}T00:02:00"),
    ("Campaign", "Chronic Offender", "Chronic Offender", "Campaign status", "In budget", "Out of budget", None, None, f"{DATE}T09:30:00"),
    # Flaps in the evening.
    ("Campaign", "Evening Cap", "Evening Cap", "Campaign daily budget", "$120.00", "$120.00", 120, 120, f"{DATE}T00:01:00"),
    ("Campaign", "Evening Cap", "Evening Cap", "Campaign status", "Out of budget", "In budget", None, None, f"{DATE}T00:03:00"),
    ("Campaign", "Evening Cap", "Evening Cap", "Campaign status", "In budget", "Out of budget", None, None, f"{DATE}T18:00:00"),
    ("Campaign", "Evening Cap", "Evening Cap", "Campaign status", "Out of budget", "In budget", None, None, f"{DATE}T19:00:00"),
    ("Campaign", "Evening Cap", "Evening Cap", "Campaign status", "In budget", "Out of budget", None, None, f"{DATE}T21:00:00"),
    # Healthy, with a human budget edit so "last action" has something to find.
    ("Campaign", "Healthy One", "Healthy One", "Campaign status", "Out of budget", "In budget", None, None, f"{DATE}T00:04:00"),
    ("Campaign", "Healthy One", "Healthy One", "Campaign daily budget", "$80.00", "$95.00", 80, 95, f"{DATE}T11:15:00"),
    # Paused for the afternoon.
    ("Campaign", "Half Paused", "Half Paused", "Campaign status", "Out of budget", "In budget", None, None, f"{DATE}T00:05:00"),
    ("Campaign", "Half Paused", "Half Paused", "Campaign status", "Delivering", "Paused", None, None, f"{DATE}T12:00:00"),
    # Account level, no campaign: a row the reader is meant to drop.
    ("Account", "", "", "Campaign status", "In budget", "Out of budget", None, None, f"{DATE}T13:00:00"),
]

CAMPAIGNS = 4          # campaign-days the pipeline should score


def build(out: Path) -> Path:
    wb = openpyxl.Workbook()
    history = wb.active
    history.title = "History"
    history.append(HEADER)
    for row in ROWS:
        history.append(list(row))

    meta = wb.create_sheet("Extraction Metadata")
    for pair in (("Account", "Synthetic Test Account"),
                 ("Marketplace", "United States"),
                 ("Date range", f"{DATE} to {DATE}"),
                 ("Extraction run ID", "synthetic-0001"),
                 ("Status", "Complete"),
                 ("Rows expected", len(ROWS)),
                 ("Rows exported", len(ROWS)),
                 ("Duplicate rows skipped", 0),
                 ("Pages processed", 1)):
        meta.append(list(pair))

    summary = wb.create_sheet("Summary Metrics")
    summary.append(["Metric", "Value"])
    for pair in (("Spend", 1500.0), ("Sales", 6000.0), ("ROAS", 4.0),
                 ("Impressions", 250000)):
        summary.append(list(pair))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
