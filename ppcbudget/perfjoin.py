"""Optional join against a campaign performance report.

The change-history export carries no per-campaign spend, so dollar figures are
limited to the ~9% of campaigns whose budget happens to appear in a budget
change row. Dropping in any Amazon Ads campaign report that has Campaign,
Spend, Sales and Budget columns lifts that to full coverage.

Join coverage is reported in both directions -- a silent join failure is how
dashboards start lying.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from .scoring import CampaignDay

# Header aliases seen across Amazon Ads report variants.
ALIASES = {
    "campaign": ["campaign", "campaign name", "campaigns"],
    "spend": ["spend", "cost", "total spend"],
    "sales": ["sales", "total sales", "14 day total sales", "7 day total sales",
              "attributed sales", "total advertising cost of sales"],
    "budget": ["budget", "daily budget", "campaign daily budget"],
    "impressions": ["impressions", "impr"],
    "clicks": ["clicks"],
    "orders": ["orders", "total orders", "14 day total orders", "7 day total orders"],
    "roas": ["roas", "total roas", "return on ad spend"],
}


@dataclass(slots=True)
class PerfRecord:
    campaign_raw: str
    spend: float | None = None
    sales: float | None = None
    budget: float | None = None
    impressions: float | None = None
    clicks: float | None = None
    orders: float | None = None
    roas: float | None = None


@dataclass(slots=True)
class JoinReport:
    path: Path
    rows_read: int = 0
    matched: int = 0
    unmatched_perf: list[str] = field(default_factory=list)
    unmatched_history: list[str] = field(default_factory=list)
    budgets_added: int = 0
    roas_added: int = 0

    @property
    def coverage(self) -> float:
        total = self.matched + len(self.unmatched_history)
        return self.matched / total if total else 0.0


def campaign_key(name: str) -> str:
    """Normalize for joining: casefold, collapse whitespace, strip punctuation runs."""
    return re.sub(r"\s+", " ", str(name).strip()).casefold()


def _num(value) -> float | None:
    if value is None or value == "":
        return None
    text = re.sub(r"[^\d.\-]", "", str(value))
    if text in ("", "-", ".", "-."):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _map_headers(header: list) -> dict[str, int]:
    lookup = {}
    normalized = [re.sub(r"\s+", " ", str(h or "").strip().lower()) for h in header]
    for field_name, names in ALIASES.items():
        for i, h in enumerate(normalized):
            if h in names:
                lookup[field_name] = i
                break
    return lookup


def _read_rows(path: Path) -> tuple[list, list[list]]:
    if path.suffix.lower() in (".csv", ".tsv", ".txt"):
        delim = "\t" if path.suffix.lower() == ".tsv" else ","
        with path.open(newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.reader(fh, delimiter=delim))
        return (rows[0], rows[1:]) if rows else ([], [])

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    # Some Amazon reports carry a title block before the real header.
    for i, row in enumerate(rows[:10]):
        if row and _map_headers(list(row)).get("campaign") is not None:
            return list(row), [list(r) for r in rows[i + 1:]]
    return list(rows[0]), [list(r) for r in rows[1:]]


def load_performance(path: str | Path) -> tuple[dict[str, PerfRecord], JoinReport]:
    path = Path(path)
    header, rows = _read_rows(path)
    cols = _map_headers(header)
    if "campaign" not in cols:
        raise ValueError(
            f"{path.name}: no Campaign column found. Expected one of: "
            + ", ".join(ALIASES['campaign'])
        )

    report = JoinReport(path=path)
    records: dict[str, PerfRecord] = {}

    def at(row, name):
        i = cols.get(name)
        return row[i] if i is not None and i < len(row) else None

    for row in rows:
        if not row or not at(row, "campaign"):
            continue
        raw = str(at(row, "campaign")).strip()
        key = campaign_key(raw)
        report.rows_read += 1
        rec = records.get(key) or PerfRecord(campaign_raw=raw)
        for f in ("spend", "sales", "budget", "impressions", "clicks", "orders", "roas"):
            v = _num(at(row, f))
            if v is not None:
                prev = getattr(rec, f)
                # Reports can be split by day/placement; sum the additive ones.
                setattr(rec, f, v if prev is None or f in ("budget", "roas") else prev + v)
        records[key] = rec

    for rec in records.values():
        if rec.roas is None and rec.spend and rec.sales is not None and rec.spend > 0:
            rec.roas = rec.sales / rec.spend
    return records, report


def apply_to(days: list[CampaignDay], records: dict[str, PerfRecord],
             report: JoinReport) -> None:
    """Overlay observed budgets onto scored days; unmatched names are reported."""
    seen: set[str] = set()
    for day in days:
        key = campaign_key(day.campaign)
        rec = records.get(key)
        if rec is None:
            report.unmatched_history.append(day.campaign)
            continue
        seen.add(key)
        report.matched += 1
        if rec.budget is not None and rec.budget > 0:
            if day.budget.source == "unknown":
                report.budgets_added += 1
            day.budget.value = rec.budget
            day.budget.time_weighted = rec.budget
            day.budget.source = "perf_report"
        if rec.roas is not None:
            report.roas_added += 1
        day.perf = rec  # type: ignore[attr-defined]

    report.unmatched_perf = [r.campaign_raw for k, r in records.items() if k not in seen]
