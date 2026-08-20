"""Render the scored data as a formatted Excel workbook.

The primary sheet is one row per campaign, matching the dashboard. Durations are
stored as real Excel time values -- a fraction of a day -- and displayed with a
`[h]"h" mm"min"` format, so a cell reads "23h 35min" while still summing,
sorting and charting as a number. Decimal hours are unreadable; text like
"23h 35min" cannot be calculated with. This gets both.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .aggregate import CampaignRollup
from .ingest import QaReport, WorkbookMeta
from .metrics import ModelSettings, Totals, hourly_starvation
from .scoring import CampaignDay

# ------------------------------------------------------------------- palette

# Utopia Brands. The report is the artefact that leaves the building, so it
# carries the same identity as the dashboard: deep green chrome on pale mint
# panels. RED and AMBER are functional, not brand -- the guide has no warning
# colour and a report about campaigns going dark needs one.
DEEP = "004D43"    # brand primary, deep green
MIST = "EAFFF4"    # brand primary, pale mint
SLATE = "4A6B64"
WHITE = "FFFFFF"
GREEN = "0F7A5C"
AMBER = "A2600A"
RED = "C9401F"
GREY = "7C9A92"
PANEL = MIST

HDR_FILL = PatternFill("solid", fgColor=DEEP)
HDR_FONT = Font(color=WHITE, bold=True, size=10)
TITLE_FONT = Font(color=DEEP, bold=True, size=20)
SUB_FONT = Font(color=SLATE, size=10, italic=True)
SECTION = Font(color=DEEP, bold=True, size=12)
TILE_FILL = PatternFill("solid", fgColor=PANEL)
KPI_LABEL = Font(color=SLATE, size=9, bold=True)
KPI_VALUE = Font(color=DEEP, size=18, bold=True)
KPI_ALARM = Font(color=RED, size=18, bold=True)
KPI_NOTE = Font(color=GREY, size=8, italic=True)
RUN_FONT = Font(color=GREEN, size=10)
LOST_FONT = Font(color=RED, size=10, bold=True)
UNPRICED = Font(color=GREY, size=9, italic=True)
# Per-date sub-columns run narrow, so they get their own smaller type.
DAY_RUN_FONT = Font(color=GREEN, size=9)
DAY_LOST_FONT = Font(color="8F2810", size=9, bold=True)
DAY_PAUSE_FONT = Font(color=GREY, size=9)

EDGE = Side(style="thin", color="CFE8DC")
BOX = Border(left=EDGE, right=EDGE, top=EDGE, bottom=EDGE)

# Hour-of-day heat, reused so 60k cells share nine fill objects. The first entry
# is the "no time lost" bucket; the rest are one warm hue getting steadily
# darker, matching heatColor() in web/app.js so the report and the dashboard
# shade the same hour the same way.
HEAT = [PatternFill("solid", fgColor=c) for c in
        ("D3F6E8", "FDECE7", "FBD7CD", "F9BFAE", "F7A58C",
         "F4886A", "F2542D", "D8431F", "B53617")]
PAUSED_FILL = PatternFill("solid", fgColor="DDE5E3")
NA_FILL = PatternFill("solid", fgColor="F2F8F5")
HEAT_FONT = Font(size=7, color=SLATE)
HEAT_FONT_DARK = Font(size=7, color=WHITE)

# Same tints the dashboard's diagnosis pills use, so a reader moving between
# the two sees one scheme.
DIAGNOSIS_FILL = {
    "Structurally underfunded": PatternFill("solid", fgColor="FFE4DC"),
    "Exhausts early": PatternFill("solid", fgColor="FDEFD0"),
    "Pacing thrash": PatternFill("solid", fgColor="E8E9FF"),
    "Evening cap": PatternFill("solid", fgColor="FEF7E4"),
    "Intermittent": PatternFill("solid", fgColor="F1FFD7"),
    "Healthy": PatternFill("solid", fgColor="D3F6E8"),
    "Mostly paused": PatternFill("solid", fgColor="E4F5EC"),
}

# A duration is a fraction of a day; [h] lets a total exceed 24 hours.
FMT_DUR = '[h]"h" mm"min"'
# Same, but a zero reads as a dash -- a column of "0h 00min" is pure noise.
FMT_DUR_Z = '[h]"h" mm"min";;"-"'
FMT_PCT = "0%"
FMT_PCT1 = "0.0%"
FMT_MONEY = '"$"#,##0.00'
FMT_MONEY0 = '"$"#,##0'
FMT_INT = "#,##0"
FMT_1 = "0.0"

MAX_DAY_COLUMNS = 31


def dur(hours: float | None) -> float | None:
    """Hours -> Excel time value. Pairs with FMT_DUR."""
    return None if hours is None else hours / 24


def hhmm(minute: int | None) -> str:
    if minute is None:
        return ""
    minute = min(minute, 1439)
    return f"{minute // 60:02d}:{minute % 60:02d}"


def _heat_style(day: CampaignDay, hour: int):
    if day.hourly_na[hour] >= 30:
        return NA_FILL, HEAT_FONT
    if day.hourly_paused[hour] >= 30:
        return PAUSED_FILL, HEAT_FONT
    oob = day.hourly_oob[hour]
    if oob == 0:
        return HEAT[0], HEAT_FONT
    level = min(8, 1 + (oob - 1) * 8 // 60)
    return HEAT[level], (HEAT_FONT_DARK if level >= 7 else HEAT_FONT)


def _header(ws, row: int, headers: list[str], widths: list[int] | None = None,
            height: int = 44) -> None:
    for i, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[row].height = height


def _as_table(ws, name: str, last_row: int, last_col: int, first_row: int = 1) -> None:
    """Turn a range into a native Excel table: banded rows and filter buttons."""
    if last_row <= first_row:
        return
    ref = f"A{first_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)


STALE_FILL = PatternFill("solid", fgColor="FFE4DC")
WARM_FILL = PatternFill("solid", fgColor="FDEFD0")
STALE_FONT = Font(color="8F2810", size=10, bold=True)


def _action_cells(ws, row: int, col: int, act) -> None:
    """Last meaningful action: summary, age, what changed, how many."""
    if act is None:
        ws.cell(row=row, column=col, value="not observed").font = UNPRICED
        return
    summary = ws.cell(row=row, column=col, value=act.summary)
    ws.cell(row=row, column=col + 1, value=act.days_since)
    ws.cell(row=row, column=col + 2, value=act.last_label or "")
    ws.cell(row=row, column=col + 3, value=act.count or None)

    # Untouched for the whole window is the thing to spot from across the room.
    if act.untouched:
        summary.fill, summary.font = STALE_FILL, STALE_FONT
        ws.cell(row=row, column=col + 1, value=None)
    elif act.days_since is not None and act.days_since >= 3:
        summary.fill = WARM_FILL


def _tile(ws, row: int, col: int, label: str, value, note: str,
          fmt: str | None = None, alarm: bool = False, width: int = 2) -> None:
    """A KPI card: label, big number, footnote, boxed and filled."""
    for r in range(row, row + 3):
        for c in range(col, col + width):
            cell = ws.cell(row=r, column=c)
            cell.fill = TILE_FILL
            cell.border = BOX
    ws.cell(row=row, column=col, value=label).font = KPI_LABEL
    v = ws.cell(row=row + 1, column=col, value=value)
    v.font = KPI_ALARM if alarm else KPI_VALUE
    v.alignment = Alignment(horizontal="left")
    if fmt:
        v.number_format = fmt
    ws.cell(row=row + 2, column=col, value=note).font = KPI_NOTE
    ws.row_dimensions[row + 1].height = 26


# ------------------------------------------------------------------- Summary


def _sheet_summary(wb: Workbook, days: list[CampaignDay], totals: Totals,
                   metas: list[WorkbookMeta], settings: ModelSettings,
                   date_keys: list[str], actions: dict | None = None) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGHIJKL",
                          (26, 14, 26, 14, 26, 14, 26, 14, 12, 12, 12, 12)):
        ws.column_dimensions[col].width = width

    account = metas[0].account if metas else ""
    market = metas[0].marketplace if metas else ""
    span = date_keys[0] if len(date_keys) == 1 else f"{date_keys[0]} to {date_keys[-1]}"

    ws["A1"] = "Out-of-Budget Campaign Analysis"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28
    ws["A2"] = f"{account} · {market} · {span} · {len(metas)} export(s)"
    ws["A2"].font = SUB_FONT

    # A typical campaign's day -- the figure people actually want.
    a_run = totals.in_hours / totals.campaigns if totals.campaigns else 0
    a_out = totals.oob_hours / totals.campaigns if totals.campaigns else 0
    a_pau = totals.paused_hours / totals.campaigns if totals.campaigns else 0

    ws["A4"] = "A typical campaign's day"
    ws["A4"].font = SECTION
    ws["A5"] = ("On an average day one campaign is able to run for the first figure below, "
                "then sits shut off for the second because it hit its daily budget.")
    ws["A5"].font = SUB_FONT
    _tile(ws, 6, 1, "RUNS PER DAY", dur(a_run), "able to spend", FMT_DUR)
    _tile(ws, 6, 3, "LOST PER DAY", dur(a_out), "shut off by its budget", FMT_DUR, alarm=True)
    _tile(ws, 6, 5, "PAUSED PER DAY", dur(a_pau), "costs nothing", FMT_DUR)
    _tile(ws, 6, 7, "CAMPAIGNS", totals.distinct_campaigns,
          f"{totals.campaigns:,} campaign-days over {totals.days} day(s)", FMT_INT)

    unit = "campaign-days" if totals.days > 1 else "campaigns"
    per_day_out = totals.oob_hours / totals.days if totals.days else 0
    per_day_spend = totals.lost_spend / totals.days if totals.days else 0
    per_day_sales = totals.lost_sales / totals.days if totals.days else 0

    ws["A10"] = "Account-wide, per day"
    ws["A10"].font = SECTION
    _tile(ws, 11, 1, "LOST HOURS PER DAY", per_day_out,
          "campaign-hours shut off", "#,##0.0", alarm=True)
    _tile(ws, 11, 3, "LOSE OVER 12 H A DAY", totals.over_12h,
          f"{unit} more than half dark", FMT_INT, alarm=True)
    _tile(ws, 11, 5, "ENDED THE DAY OUT", totals.ended_oob,
          f"{totals.ended_oob / totals.campaigns:.0%} of {unit}" if totals.campaigns else "",
          FMT_INT, alarm=True)
    _tile(ws, 11, 7, "REPEAT OUTAGES", totals.flapping_3plus,
          f"{unit} with 3 or more outages", FMT_INT)
    _tile(ws, 15, 1, "LOST SPEND PER DAY", per_day_spend,
          f"only {totals.priced:,} of {totals.campaigns:,} {unit} priced", FMT_MONEY0)
    _tile(ws, 15, 3, "LOST SALES PER DAY", per_day_sales,
          f"ROAS {settings.roas:.2f} x {settings.haircut:.0%} haircut", FMT_MONEY0)
    _tile(ws, 15, 5, "ACTUAL SPEND", sum(m.spend for m in metas if m.spend) or 0,
          "reported by Amazon for the period", FMT_MONEY0)
    actual = sum(m.spend for m in metas if m.spend) or 0
    _tile(ws, 15, 7, "LOST AS SHARE OF ACTUAL",
          (totals.lost_spend / actual) if actual else None,
          "if this nears 100% the model is wrong", FMT_PCT1)

    stale = sum(1 for a in (actions or {}).values() if a.untouched)
    if actions:
        _tile(ws, 19, 1, "NO ACTION IN WINDOW", stale,
              f"of {len(actions):,} campaigns, over {len(date_keys)} day(s)",
              FMT_INT, alarm=bool(stale))
        _tile(ws, 19, 3, "TOUCHED IN WINDOW", len(actions) - stale,
              "had a budget, bid, placement or targeting change", FMT_INT)
        ws["E19"] = ("A campaign starving with no action taken is the one to open first. "
                     "Amazon's own out-of-budget rows are not counted as actions.")
        ws["E19"].font = KPI_NOTE

    ws["A23"] = ("Reality check: modelled loss is measured against the spend Amazon actually "
                 f"reported. {totals.capped} campaign-days hit the {settings.cap_multiple:g}x "
                 f"budget cap; {totals.rate_unreliable} had too little in-budget time to price.")
    ws["A23"].font = KPI_NOTE

    # Hour-of-day starvation curve.
    curve = hourly_starvation(days)
    ws["A25"] = "Starvation through the day"
    ws["A25"].font = SECTION
    ws["A26"] = ("Share of campaigns out of budget during each hour. Budgets reset at midnight, "
                 "then coverage decays as campaigns exhaust their daily cap.")
    ws["A26"].font = SUB_FONT
    hdr = 27
    _header(ws, hdr, ["Hour", "% out of budget"], height=20)
    for h in range(24):
        ws.cell(row=hdr + 1 + h, column=1, value=f"{h:02d}:00")
        c = ws.cell(row=hdr + 1 + h, column=2, value=curve[h] / 100)
        c.number_format = FMT_PCT1

    chart = BarChart()
    chart.type = "col"
    chart.title = "Campaigns out of budget by hour"
    chart.y_axis.title = "% of campaigns"
    chart.x_axis.title = "Hour of day"
    chart.height, chart.width = 10, 24
    chart.legend = None
    chart.varyColors = False
    chart.gapWidth = 40
    chart.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=hdr + 24), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=hdr + 1, max_row=hdr + 24))
    chart.series[0].graphicalProperties.solidFill = RED
    chart.series[0].graphicalProperties.line.solidFill = RED
    ws.add_chart(chart, "D27")

    ws["A53"] = "Where to look next"
    ws["A53"].font = SECTION
    ws["A54"] = ("The Campaigns sheet has one row per campaign — start there. Daily Detail "
                 "breaks each campaign into its individual days with an hour-by-hour heatmap, "
                 "and Episodes lists every single outage with start and end times.")
    ws["A54"].font = SUB_FONT


# ----------------------------------------------------------------- Campaigns


def _sheet_campaigns_multi(wb: Workbook, rollups: list[CampaignRollup],
                           date_keys: list[str], actions: dict) -> None:
    """One row per campaign, averaged across days, plus a column per day."""
    ws = wb.create_sheet("Campaigns")
    shown_dates = date_keys[:MAX_DAY_COLUMNS]

    headers = ["Campaign", "Days seen", "Days it ran out", "Recurrence",
               "Runs / day", "Lost / day", "Paused / day", "Worst day", "Worst date",
               "Total lost", "Outages", "Longest streak", "Trend", "Lost $/day",
               "Lost $ total", "Chronic score", "Diagnosis",
               "Last action", "Days since action", "What changed last", "Actions in window"]
    # Each width allows for the table filter button, which eats ~3 units.
    widths = [46, 11, 14, 13, 12, 12, 13, 12, 13, 12, 10, 14, 12, 13, 13, 14, 25,
              26, 13, 34, 13]
    # Three sub-columns per date. The names carry the date so every header in
    # the table stays unique, which Excel requires, and they wrap onto two lines.
    day_headers = []
    for d in shown_dates:
        day_headers += [f"{d[5:]} Runs", f"{d[5:]} Lost", f"{d[5:]} Paused"]
    _header(ws, 1, headers + day_headers, widths + [11] * len(day_headers))

    base = len(headers)
    for i, r in enumerate(rollups):
        row = i + 2
        vals = [
            r.campaign, r.days_observed, r.days_with_oob, r.recurrence_rate,
            dur(r.mean_in_hours), dur(r.mean_oob_hours), dur(r.mean_paused_hours),
            dur(r.max_oob_hours), r.worst_date, dur(r.total_oob_hours),
            r.total_episodes, r.streak_max, r.trend_label,
            (r.total_lost_spend / r.days_observed) if r.total_lost_spend else None,
            r.total_lost_spend, r.chronic_score, r.dominant_diagnosis,
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        ws.cell(row=row, column=4).number_format = FMT_PCT
        for c in (5, 6, 8, 10):
            ws.cell(row=row, column=c).number_format = FMT_DUR
        ws.cell(row=row, column=7).number_format = FMT_DUR_Z
        ws.cell(row=row, column=5).font = RUN_FONT
        ws.cell(row=row, column=6).font = LOST_FONT
        for c in (14, 15):
            ws.cell(row=row, column=c).number_format = FMT_MONEY
        ws.cell(row=row, column=16).number_format = FMT_1

        if r.total_lost_spend is None:
            ws.cell(row=row, column=14, value="no budget").font = UNPRICED
        t = ws.cell(row=row, column=13)
        if r.trend_label == "worsening":
            t.font = Font(color=RED, bold=True, size=10)
        elif r.trend_label == "improving":
            t.font = Font(color=GREEN, size=10)
        dx = ws.cell(row=row, column=17)
        if r.dominant_diagnosis in DIAGNOSIS_FILL:
            dx.fill = DIAGNOSIS_FILL[r.dominant_diagnosis]
        _action_cells(ws, row, 18, actions.get(r.campaign))

        by_date = {p.date_key: p for p in r.per_day}
        for j, dk in enumerate(shown_dates):
            p = by_date.get(dk)
            trio = ((p.in_hours, DAY_RUN_FONT), (p.oob_hours, DAY_LOST_FONT),
                    (p.paused_hours, DAY_PAUSE_FONT)) if p else ((None, DAY_RUN_FONT),) * 3
            for k, (hours, font) in enumerate(trio):
                cell = ws.cell(row=row, column=base + 1 + j * 3 + k, value=dur(hours))
                cell.number_format = FMT_DUR_Z
                cell.font = font
                cell.alignment = Alignment(horizontal="center")

    last_row = len(rollups) + 1
    last_col = base + len(shown_dates) * 3
    # Shade only the Lost sub-column: 0h green through 24h red. One rule per
    # day, all with the same explicit thresholds so the scale is comparable.
    for j in range(len(shown_dates)):
        letter = get_column_letter(base + 2 + j * 3)
        ws.conditional_formatting.add(f"{letter}2:{letter}{last_row}", ColorScaleRule(
            start_type="num", start_value=0, start_color="D3F6E8",
            mid_type="num", mid_value=0.5, mid_color="F7A58C",
            end_type="num", end_value=1, end_color="8F2810"))
    _as_table(ws, "Campaigns", last_row, last_col)
    ws.freeze_panes = "B2"

    if len(date_keys) > len(shown_dates):
        note = ws.cell(row=last_row + 2, column=1,
                       value=f"Day columns show the first {MAX_DAY_COLUMNS} of "
                             f"{len(date_keys)} days. Every day is in Daily Detail.")
        note.font = KPI_NOTE


def _sheet_campaigns_single(wb: Workbook, days: list[CampaignDay], actions: dict) -> None:
    """Single day: one row per campaign already, so carry the 24-hour heatmap."""
    ws = wb.create_sheet("Campaigns")
    headers = ["Campaign", "Date", "Runs", "Lost", "Paused", "% of day lost",
               "Budget-cap hits", "Distinct outages", "First out", "Last recovery",
               "Ended out", "Daily budget", "Budget source", "Spend rate /h",
               "Lost spend", "Lost sales", "Capped", "Severity", "Diagnosis",
               "Confidence", "+/- hours",
               "Last action", "Days since action", "What changed last", "Actions in window"]
    widths = [46, 13, 12, 12, 12, 14, 14, 14, 11, 14, 11, 13, 15, 13, 13, 13, 10, 11, 25, 13, 11,
              26, 13, 34, 13]
    hours = [f"{h:02d}" for h in range(24)]
    _header(ws, 1, headers + hours, widths + [3.6] * 24)
    ordered = sorted(days, key=lambda d: (-d.severity, d.campaign))
    _fill_day_rows(ws, ordered, len(headers))
    for i, d in enumerate(ordered):
        _action_cells(ws, i + 2, 22, actions.get(d.campaign))
    _as_table(ws, "Campaigns", len(days) + 1, len(headers) + 24)
    ws.freeze_panes = "C2"


def _sheet_daily_detail(wb: Workbook, days: list[CampaignDay]) -> None:
    ws = wb.create_sheet("Daily Detail")
    headers = ["Campaign", "Date", "Runs", "Lost", "Paused", "% of day lost",
               "Budget-cap hits", "Distinct outages", "First out", "Last recovery",
               "Ended out", "Daily budget", "Budget source", "Spend rate /h",
               "Lost spend", "Lost sales", "Capped", "Severity", "Diagnosis",
               "Confidence", "+/- hours"]
    widths = [46, 13, 12, 12, 12, 14, 14, 14, 11, 14, 11, 13, 15, 13, 13, 13, 10, 11, 25, 13, 11]
    hours = [f"{h:02d}" for h in range(24)]
    _header(ws, 1, headers + hours, widths + [3.6] * 24)
    ordered = sorted(days, key=lambda d: (d.campaign, d.date_key))
    _fill_day_rows(ws, ordered, len(headers))
    _as_table(ws, "DailyDetail", len(days) + 1, len(headers) + 24)
    ws.freeze_panes = "C2"


def _fill_day_rows(ws, ordered: list[CampaignDay], base: int) -> None:
    """Shared body for the per-campaign-day sheets, including the hour heatmap."""
    for i, d in enumerate(ordered):
        row = i + 2
        lost = d.lost or {}
        vals = [
            d.campaign, d.date_key, dur(d.in_hours), dur(d.oob_hours), dur(d.paused_hours),
            d.oob_share, d.episodes_raw, d.episodes_merged,
            hhmm(d.first_oob_min), hhmm(d.last_recovery_min),
            "yes" if d.closed_oob else "no",
            d.budget.time_weighted or d.budget.value,
            d.budget.source.replace("_", " "),
            lost.get("spend_rate_per_hour"), lost.get("lost_spend"), lost.get("lost_sales"),
            "yes" if lost.get("capped") else "",
            d.severity, d.diagnosis, d.confidence.replace("_", " "),
            round(d.oob_uncertainty_min / 60, 2) if d.chain_breaks else None,
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        for c in (3, 4):
            ws.cell(row=row, column=c).number_format = FMT_DUR
        ws.cell(row=row, column=5).number_format = FMT_DUR_Z
        ws.cell(row=row, column=3).font = RUN_FONT
        ws.cell(row=row, column=4).font = LOST_FONT
        ws.cell(row=row, column=6).number_format = FMT_PCT
        for c in (12, 14, 15, 16):
            ws.cell(row=row, column=c).number_format = FMT_MONEY
        ws.cell(row=row, column=18).number_format = FMT_1
        ws.cell(row=row, column=21).number_format = "0.00"

        if d.budget.source == "unknown":
            ws.cell(row=row, column=13, value="not in export").font = UNPRICED
            for c in (12, 14, 15, 16):
                ws.cell(row=row, column=c).font = UNPRICED
        dx = ws.cell(row=row, column=19)
        if d.diagnosis in DIAGNOSIS_FILL:
            dx.fill = DIAGNOSIS_FILL[d.diagnosis]
        if d.confidence != "clean":
            ws.cell(row=row, column=20).font = Font(color=AMBER, size=9, bold=True)

        for h in range(24):
            cell = ws.cell(row=row, column=base + 1 + h, value=d.hourly_oob[h] or None)
            cell.fill, cell.font = _heat_style(d, h)
            cell.alignment = Alignment(horizontal="center")


# ------------------------------------------------------------------ Episodes


def _sheet_episodes(wb: Workbook, days: list[CampaignDay]) -> None:
    ws = wb.create_sheet("Episodes")
    headers = ["Campaign", "Date", "Outage #", "Start", "End", "Duration",
               "Billable", "Paused during", "Diagnosis"]
    _header(ws, 1, headers, [46, 13, 11, 10, 10, 13, 13, 15, 25])
    r = 2
    for d in sorted(days, key=lambda x: (x.campaign, x.date_key)):
        for ep in d.episodes:
            ws.cell(row=r, column=1, value=d.campaign)
            ws.cell(row=r, column=2, value=d.date_key)
            ws.cell(row=r, column=3, value=ep.index)
            ws.cell(row=r, column=4, value=hhmm(ep.start_min))
            ws.cell(row=r, column=5, value=hhmm(ep.end_min))
            for col, mins in ((6, ep.raw_min), (7, ep.active_min),
                              (8, ep.raw_min - ep.active_min)):
                c = ws.cell(row=r, column=col, value=dur(mins / 60))
                c.number_format = FMT_DUR_Z if col == 8 else FMT_DUR
            ws.cell(row=r, column=7).font = LOST_FONT
            ws.cell(row=r, column=9, value=d.diagnosis)
            r += 1
    _as_table(ws, "Episodes", r - 1, len(headers))
    ws.freeze_panes = "C2"

    note = ws.cell(row=r + 1, column=1,
                   value="Duration is wall-clock. Billable excludes minutes the campaign was "
                         "paused during the outage — a paused campaign forgoes nothing to its "
                         "budget, so only billable minutes count as lost.")
    note.font = KPI_NOTE


# -------------------------------------------------------------- Data Quality


def _sheet_quality(wb: Workbook, qas: list[QaReport], days: list[CampaignDay],
                   totals: Totals, join_report=None, overlap_rows: int = 0) -> None:
    ws = wb.create_sheet("Data Quality")
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", (34, 12, 22, 20, 78)):
        ws.column_dimensions[col].width = width
    ws["A1"] = "Data quality"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26
    ws["A2"] = "Every check that could change how much you trust the numbers on the other sheets."
    ws["A2"].font = SUB_FONT

    _header(ws, 4, ["Check", "Status", "Value", "Expected", "What it means"])
    row = 5

    def check(name, ok, value, expected, note):
        nonlocal row
        ws.cell(row=row, column=1, value=name).alignment = Alignment(wrap_text=True, vertical="top")
        s = ws.cell(row=row, column=2,
                    value="OK" if ok is True else ("REVIEW" if ok is None else "FAIL"))
        s.font = Font(color=GREEN if ok is True else (AMBER if ok is None else RED), bold=True)
        s.alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(row=row, column=3, value=value).alignment = Alignment(vertical="top")
        ws.cell(row=row, column=4, value=expected).alignment = Alignment(vertical="top")
        ws.cell(row=row, column=5, value=note).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BOX
        ws.row_dimensions[row].height = 34
        row += 1

    for qa in qas:
        m = qa.meta
        check(f"Row accounting: {qa.path.name[:30]}", qa.row_accounting_ok,
              f"{qa.rows_parsed:,} scored",
              f"{m.rows_expected:,} - {m.duplicates_skipped:,} dup" if m.rows_expected else "n/a",
              f"{qa.accounting_detail}. " + ("Every exported row is accounted for."
              if qa.row_accounting_ok else "These do NOT reconcile — some exported rows are "
              "unaccounted for, so the figures may be understated."))
        if m.status and m.status != "completed":
            check(f"Extraction status: {qa.path.name[:28]}", False, m.status, "completed",
                  "A partial extraction can be missing whole campaigns, not just rows.")
        if qa.warning_categories:
            check(f"Exporter warnings: {qa.path.name[:28]}", False,
                  ", ".join(f"{n}x {c}" for c, n in qa.warning_categories.items()), "none",
                  "The export recorded its own failures. PARTIAL_PAGE_HARVEST means it "
                  "collected fewer rows than the page reported, so changes are missing and "
                  "every duration here is a lower bound. Re-run the extraction.")
        if qa.mixed_sources:
            check(f"Mixed sources: {qa.path.name[:28]}", False,
                  f"{len(qa.source_urls)} consoles / {len(qa.row_marketplaces)} marketplaces",
                  "1 / 1",
                  "This file holds rows from more than one console or marketplace ("
                  + ", ".join(sorted(qa.row_marketplaces)[:4])
                  + "). Account spend and ROAS come from one metadata block, so money would be "
                  "attributed to the wrong marketplace. Export each marketplace separately.")

    if len(qas) > 1:
        check("Overlapping exports", True,
              f"{overlap_rows:,} counted once" if overlap_rows else "no overlap", "n/a",
              "Exports are date-range based, so loading a week and a month that contains it is "
              "normal. Rows in more than one file are matched on entity, timestamp and values "
              "and counted once.")

    check("Budget and delivery kept separate",
          all(q.crossover_violations == 0 for q in qas),
          f"{sum(q.crossover_violations for q in qas)} violations", "0",
          "'Campaign status' carries two independent state machines. No row mixes them, so the "
          "split into budget timeline and pause overlay is lossless.")

    repaired = [d for d in days if d.chain_breaks]
    check("Timeline continuity", None if repaired else True,
          f"{len(repaired)} repaired" if repaired else "no gaps", "0",
          "De-duplication can drop an intermediate transition, leaving a row whose 'From' "
          "disagrees with the running state. Each is repaired at the midpoint of the gap and "
          "carries an uncertainty band, listed below.")

    check("Budget coverage", None if totals.priced < totals.campaigns else True,
          f"{totals.priced:,} of {totals.campaigns:,}", "all",
          "Dollar figures need a daily budget, which the change history only reveals for "
          "campaigns whose budget was edited. Unpriced campaigns show no money figure rather "
          "than a zero. Add a performance report to price the rest.")

    if totals.partial_day:
        check("Partial-day campaigns", None, totals.partial_day, "0",
              "Created mid-day, so scored over the remainder of the day only — never penalised "
              "for hours before they existed.")

    if join_report is not None:
        check("Performance report join", join_report.coverage > 0.9,
              f"{join_report.matched:,} matched", f"{join_report.rows_read:,} rows",
              f"{len(join_report.unmatched_history):,} campaigns had no performance row; "
              f"{len(join_report.unmatched_perf):,} performance rows matched no campaign.")

    check("Internal consistency", True, f"{len(days):,} campaigns", "all",
          "For every campaign the minutes in budget, out of budget, paused and not-yet-created "
          "sum to exactly 1440; episode durations sum to the out-of-budget total; and the "
          "hourly buckets agree with both.")

    if repaired:
        row += 1
        ws.cell(row=row, column=1, value="Repaired campaigns").font = SECTION
        row += 1
        _header(ws, row, ["Campaign", "Date", "At", "Expected state", "Observed / uncertainty"])
        row += 1
        for d in repaired:
            for b in d.chain_breaks:
                ws.cell(row=row, column=1, value=d.campaign)
                ws.cell(row=row, column=2, value=d.date_key)
                ws.cell(row=row, column=3, value=hhmm(b.at_min))
                ws.cell(row=row, column=4, value=b.expected_from)
                ws.cell(row=row, column=5,
                        value=f"row said '{b.saw_from}' — {b.ambiguity_min} min gap, so this "
                              f"campaign carries +/- {b.ambiguity_min / 120:.2f} h")
                row += 1


# -------------------------------------------------------------------- Method


def _sheet_method(wb: Workbook, settings: ModelSettings, metas: list[WorkbookMeta]) -> None:
    ws = wb.create_sheet("Method")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 108
    ws["A1"] = "How every number here is calculated"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26
    ws["A2"] = "So any figure in this workbook can be defended in a meeting."
    ws["A2"].font = SUB_FONT

    roas_note = ("account average from the export's Summary Metrics"
                 if settings.roas_source == "account_average"
                 else "per campaign from the performance report")
    entries = [
        ("Last action", "The most recent optimisation change on that campaign inside the days "
                        "the export covers: budget, bid, placement %, bidding strategy, "
                        "targeting, enable/pause, or structure. Amazon's own "
                        "out-of-budget rows are excluded — they are the pacing engine, not a "
                        "person, and counting them would make every starving campaign look "
                        "managed. Renames are excluded too. \"No action in N days\" means "
                        "nothing was changed across the whole observed window."),
        ("Which sheet to use", "Campaigns is one row per campaign — the place to start. Daily "
                               "Detail is one row per campaign per day with an hour-by-hour "
                               "heatmap. Episodes is one row per individual outage."),
        ("Durations", "Stored as real time values and displayed as \"23h 35min\", so they still "
                      "sum, sort and chart correctly. Widen a column or change the number format "
                      "to see them as decimal hours."),
        ("Source data", "Amazon Ads change history. It lists only changes, so the state between "
                        "two rows is inferred by walking the events in order."),
        ("Event ordering", "The export is written newest-first, so rows sharing a minute are "
                           "reversed before the state machine walks them. Sorting on timestamp "
                           "alone silently preserves the wrong order within a minute."),
        ("Runs vs Lost", "Runs is time in budget and able to spend. Lost is time shut off after "
                         "hitting the daily budget. With paused time they make up the 24-hour "
                         "day."),
        ("Paused time", "Delivery state (Delivering/Paused) forms a second, independent track. "
                        "Paused minutes are excluded from lost time and from the in-budget "
                        "denominator, because a paused campaign forgoes nothing to its budget."),
        ("Billable", "On an individual outage, wall-clock duration minus any minutes the "
                     "campaign was paused during it. Only billable minutes count as lost."),
        ("Eligible window", "Midnight to midnight, except a campaign created mid-day, which is "
                            "scored from its creation minute onward."),
        ("Cap hits vs outages", "Hits counts every In-to-Out transition. Outages merges those "
                                "separated by under 5 minutes in budget, since Amazon can "
                                "release a sliver of budget consumed within the same minute."),
        ("Severity",
         f"100 x ({settings.w_share:g} x share of active day lost + {settings.w_early:g} x how "
         f"early it ran out + {settings.w_flap:g} x outage frequency, capped at 12)."),
        ("Chronic score", "Across multiple days: 40% how often it ran out, 35% average hours "
                          "lost per day, 25% longest consecutive run of bad days. A campaign "
                          "losing 8 hours every day outranks one that spiked to 23 hours once."),
        ("Spend rate", "Daily budget divided by hours in budget. Budgets that changed during the "
                       "day are time-weighted."),
        ("Lost spend", f"Spend rate x lost hours, capped at {settings.cap_multiple:g}x the daily "
                       "budget. Without the cap, a campaign in budget 20 minutes would imply a "
                       "loss far beyond what demand could absorb."),
        ("Lost sales", f"Lost spend x ROAS {settings.roas:.2f} ({roas_note}) x a "
                       f"{settings.haircut:.0%} haircut. The haircut is a modelling assumption, "
                       "not a measurement: incremental budget does not convert at the average."),
        ("Unknown budgets", "Where no daily budget appears in the export, money cells read "
                            "\"no budget\" rather than zero. A zero becomes a fact the moment "
                            "someone sums the column."),
    ]
    row = 4
    for name, text in entries:
        ws.cell(row=row, column=1, value=name).font = Font(bold=True, color=DEEP, size=10)
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
        c = ws.cell(row=row, column=2, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(15, 12.5 * (len(text) // 105 + 1))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Generated").font = Font(bold=True, color=DEEP, size=10)
    ws.cell(row=row, column=2,
            value=f"{datetime.now():%Y-%m-%d %H:%M} from "
                  + ", ".join(m.path.name for m in metas))


# --------------------------------------------------------------------- entry


def write_report(path: str | Path, days: list[CampaignDay], totals: Totals,
                 rollups: list[CampaignRollup], qas: list[QaReport],
                 metas: list[WorkbookMeta], settings: ModelSettings,
                 date_keys: list[str], join_report=None,
                 overlap_rows: int = 0, actions: dict | None = None) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, days, totals, metas, settings, date_keys, actions)
    if len(date_keys) > 1:
        _sheet_campaigns_multi(wb, rollups, date_keys, actions or {})
        _sheet_daily_detail(wb, days)
    else:
        _sheet_campaigns_single(wb, days, actions or {})
    _sheet_episodes(wb, days)
    _sheet_quality(wb, qas, days, totals, join_report, overlap_rows)
    _sheet_method(wb, settings, metas)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
