"""Read an Amazon Ads change-history export into typed records.

The export carries two independent state machines in the same `Campaign status`
change type: the budget machine (In budget / Out of budget) and the delivery
machine (Delivering / Paused). No row ever mixes the two vocabularies, so
partitioning on membership is lossless -- `QaReport.crossover_violations`
asserts that holds for every file we read.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl

BUDGET_STATES = ("In budget", "Out of budget")
DELIVERY_STATES = ("Delivering", "Paused")

CT_CAMPAIGN_STATUS = "Campaign status"
CT_DAILY_BUDGET = "Campaign daily budget"
CT_BUDGET_RULE = "Budget rule"
CT_CAMPAIGN_CREATED = "Campaign created"

# Budget rule cells read "Budget: $20.00 - Rule(s) active" with an en-dash.
_MONEY = re.compile(r"\$\s*([\d,]+(?:\.\d+)?)")


@dataclass(slots=True)
class Event:
    source_index: int  # position in the file; the intra-minute tie-break key
    date_key: str  # '2026-08-05'
    minute: int  # 0..1439 within date_key
    second: int  # 0..86399; only used to tell near-simultaneous rows apart
    level_type: str  # 'Campaign' | 'Ad group'
    level_name: str  # the entity that changed -- an ad group, not its campaign
    campaign: str
    change_type: str
    from_val: str
    to_val: str
    from_num: float | None
    to_num: float | None
    machine: str  # budget | delivery | budget_amount | budget_rule | created | other


@dataclass(slots=True)
class WorkbookMeta:
    path: Path
    account: str = ""
    marketplace: str = ""
    date_range: str = ""
    run_id: str = ""
    status: str = ""
    rows_expected: int | None = None
    rows_exported: int | None = None
    duplicates_skipped: int | None = None
    pages_processed: int | None = None
    started_at: str = ""
    completed_at: str = ""
    spend: float | None = None
    sales: float | None = None
    roas: float | None = None
    impressions: float | None = None


@dataclass(slots=True)
class QaReport:
    path: Path
    meta: WorkbookMeta
    rows_parsed: int = 0
    columns: int = 0
    crossover_violations: int = 0
    rows_unparsable_time: int = 0
    rows_no_campaign: int = 0   # account- or portfolio-level rows
    rows_blank: int = 0
    # Warnings the exporter itself recorded, and provenance of the rows. A file
    # holding two marketplaces or two extraction runs was stitched together.
    extraction_warnings: list[str] = field(default_factory=list)
    warning_categories: dict = field(default_factory=dict)
    source_urls: list[str] = field(default_factory=list)
    row_marketplaces: list[str] = field(default_factory=list)
    row_run_ids: list[str] = field(default_factory=list)
    distinct_campaigns: int = 0
    campaigns_with_budget_events: int = 0
    date_keys: list[str] = field(default_factory=list)

    @property
    def mixed_sources(self) -> bool:
        """True when one file holds rows from more than one console or run."""
        return len(self.source_urls) > 1 or len(self.row_marketplaces) > 1

    @property
    def rows_seen(self) -> int:
        """Every data row in the sheet, including ones we deliberately drop."""
        return self.rows_parsed + self.rows_no_campaign + self.rows_blank

    @property
    def row_accounting_ok(self) -> bool:
        """expected - duplicates == exported, and every exported row accounted for.

        Rows without a Campaign are real rows we cannot place on a campaign
        timeline -- account- or portfolio-level changes. They are dropped on
        purpose, so they count toward reconciliation rather than against it.
        """
        m = self.meta
        if None in (m.rows_expected, m.rows_exported, m.duplicates_skipped):
            return False
        return (
            m.rows_expected - m.duplicates_skipped == m.rows_exported
            and m.rows_exported == self.rows_seen
        )

    @property
    def accounting_detail(self) -> str:
        """Plain reconciliation line, whichever way the check lands."""
        m = self.meta
        if m.rows_expected is None:
            return "the export carries no row-count metadata to reconcile against"
        parts = [f"{m.rows_exported:,} exported", f"{self.rows_parsed:,} placed on a timeline"]
        if self.rows_no_campaign:
            parts.append(f"{self.rows_no_campaign:,} with no campaign (account or "
                         f"portfolio level, not scoreable)")
        if self.rows_blank:
            parts.append(f"{self.rows_blank:,} blank")
        if self.rows_unparsable_time:
            parts.append(f"{self.rows_unparsable_time:,} with an unreadable timestamp")
        gap = m.rows_exported - self.rows_seen
        if gap:
            parts.append(f"{gap:,} UNACCOUNTED")
        return " = ".join([parts[0], " + ".join(parts[1:])])


_WHEN_FORMATS = (
    "%d %b %Y %H:%M:%S.%f", "%d %b %Y %H:%M:%S", "%d %b %Y %H:%M",
    "%m/%d/%Y %H:%M:%S.%f", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
)


def parse_when(*candidates) -> datetime | None:
    """First readable timestamp among the given cells.

    The ISO column is preferred, but some exports fill only the human-readable
    one, so both are tried before a row is written off.
    """
    for raw in candidates:
        if raw is None or raw == "":
            continue
        if isinstance(raw, datetime):
            return raw
        text = str(raw).strip()
        if not text or text == "None":
            continue
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            pass
        for fmt in _WHEN_FORMATS:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _num(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


def parse_money(*texts) -> float | None:
    """First dollar amount found across the given cells, or None."""
    for text in texts:
        if not text:
            continue
        m = _MONEY.search(str(text))
        if m:
            return float(m.group(1).replace(",", ""))
    return None


def _classify(change_type: str, from_val: str, to_val: str) -> str:
    if change_type == CT_CAMPAIGN_STATUS:
        if from_val in BUDGET_STATES or to_val in BUDGET_STATES:
            return "budget"
        if from_val in DELIVERY_STATES or to_val in DELIVERY_STATES:
            return "delivery"
        return "other"
    if change_type == CT_DAILY_BUDGET:
        return "budget_amount"
    if change_type == CT_BUDGET_RULE:
        return "budget_rule"
    if change_type == CT_CAMPAIGN_CREATED:
        return "created"
    return "other"


def _read_meta(wb, path: Path) -> WorkbookMeta:
    meta = WorkbookMeta(path=path)

    if "Extraction Metadata" in wb.sheetnames:
        pairs = {
            str(r[0]).strip(): r[1]
            for r in wb["Extraction Metadata"].iter_rows(values_only=True)
            if r and r[0]
        }
        meta.account = str(pairs.get("Account", "") or "")
        meta.marketplace = str(pairs.get("Marketplace", "") or "")
        meta.date_range = str(pairs.get("Date range", "") or "")
        meta.run_id = str(pairs.get("Extraction run ID", "") or "")
        meta.status = str(pairs.get("Status", "") or "")
        meta.started_at = str(pairs.get("Started at", "") or "")
        meta.completed_at = str(pairs.get("Completed at", "") or "")
        for attr, key in (
            ("rows_expected", "Rows expected"),
            ("rows_exported", "Rows exported"),
            ("duplicates_skipped", "Duplicate rows skipped"),
            ("pages_processed", "Pages processed"),
        ):
            v = _num(pairs.get(key))
            if v is not None:
                setattr(meta, attr, int(v))

    if "Summary Metrics" in wb.sheetnames:
        for row in wb["Summary Metrics"].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key, value = str(row[0]).strip().lower(), _num(row[1])
            if key in ("spend", "sales", "roas", "impressions"):
                setattr(meta, key, value)

    return meta


def _read_warnings(wb, qa: QaReport) -> None:
    """The exporter records its own failures on an 'Errors and Warnings' sheet.

    Ignoring it means silently inheriting whatever it could not collect, so the
    categories and a sample of messages are carried through to Data Quality.
    """
    sheet = next((n for n in wb.sheetnames if "error" in n.lower()
                  or "warning" in n.lower()), None)
    if sheet is None:
        return
    rows = list(wb[sheet].iter_rows(values_only=True))
    if len(rows) < 2:
        return
    header = [str(h).strip().lower() if h else "" for h in rows[0]]

    def idx(name):
        return header.index(name) if name in header else None

    i_cat, i_msg, i_pg = idx("category"), idx("message"), idx("page number")
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        cat = str(row[i_cat]).strip() if i_cat is not None and row[i_cat] else "WARNING"
        msg = str(row[i_msg]).strip() if i_msg is not None and row[i_msg] else ""
        page = row[i_pg] if i_pg is not None else None
        qa.warning_categories[cat] = qa.warning_categories.get(cat, 0) + 1
        if len(qa.extraction_warnings) < 8:
            qa.extraction_warnings.append(
                f"{cat}" + (f" (page {page})" if page else "") + (f": {msg}" if msg else ""))


def load_history(path: str | Path) -> tuple[list[Event], WorkbookMeta, QaReport]:
    """Parse one change-history workbook."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "History" not in wb.sheetnames:
        raise ValueError(f"{path.name}: no 'History' sheet -- is this a change-history export?")

    meta = _read_meta(wb, path)
    qa = QaReport(path=path, meta=meta)
    _read_warnings(wb, qa)

    rows = wb["History"].iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError(f"{path.name}: History sheet is empty")
    col = {str(name).strip(): i for i, name in enumerate(header) if name}
    qa.columns = len(header)

    required = ["Campaign", "Change type", "From", "To", "Date and time (ISO)"]
    missing = [c for c in required if c not in col]
    if missing:
        raise ValueError(f"{path.name}: missing expected column(s): {', '.join(missing)}")

    def cell(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    events: list[Event] = []
    campaigns: set[str] = set()
    dates: set[str] = set()

    for idx, row in enumerate(rows):
        if row is None or not any(v is not None for v in row):
            qa.rows_blank += 1
            continue
        campaign = cell(row, "Campaign")
        if not campaign:
            # Account- and portfolio-level rows have no campaign to attach to.
            qa.rows_no_campaign += 1
            continue
        campaign = str(campaign).strip()
        campaigns.add(campaign)

        when = parse_when(cell(row, "Date and time (ISO)"), cell(row, "Date and time"))
        if when is None:
            qa.rows_unparsable_time += 1
            continue

        from_val = "" if cell(row, "From") is None else str(cell(row, "From")).strip()
        to_val = "" if cell(row, "To") is None else str(cell(row, "To")).strip()
        change_type = str(cell(row, "Change type") or "").strip()

        # The partition is only lossless if no row straddles both vocabularies.
        if change_type == CT_CAMPAIGN_STATUS:
            in_b = (from_val in BUDGET_STATES, to_val in BUDGET_STATES)
            if in_b[0] != in_b[1]:
                qa.crossover_violations += 1

        for name, bucket in (("Source URL", qa.source_urls),
                             ("Marketplace", qa.row_marketplaces),
                             ("Run ID", qa.row_run_ids)):
            v = cell(row, name)
            if v and str(v).strip() not in bucket:
                bucket.append(str(v).strip())

        date_key = when.date().isoformat()
        dates.add(date_key)
        events.append(
            Event(
                source_index=idx,
                date_key=date_key,
                minute=when.hour * 60 + when.minute,
                second=when.hour * 3600 + when.minute * 60 + when.second,
                level_type=str(cell(row, "Change level type") or "").strip(),
                level_name=str(cell(row, "Change level name") or "").strip(),
                campaign=campaign,
                change_type=change_type,
                from_val=from_val,
                to_val=to_val,
                from_num=_num(cell(row, "From (numeric)")),
                to_num=_num(cell(row, "To (numeric)")),
                machine=_classify(change_type, from_val, to_val),
            )
        )

    wb.close()

    # No timestamps means no timeline, and no timeline means out-of-budget hours
    # cannot be measured at all. Say exactly that, plus whatever the exporter
    # already admitted, instead of a generic "could not be read".
    if not events and qa.rows_unparsable_time:
        detail = [
            f"{path.name}: none of the {qa.rows_unparsable_time:,} rows have a usable "
            "timestamp — 'Date and time' and 'Date and time (ISO)' are both empty. "
            "Without a time on each change there is no way to reconstruct when a "
            "campaign was in or out of budget, so hours cannot be measured."
        ]
        if qa.warning_categories:
            worst = ", ".join(f"{n}x {c}" for c, n in qa.warning_categories.items())
            detail.append(f"The export also recorded its own problems ({worst}), "
                          "so rows are missing as well.")
        if not qa.row_accounting_ok and meta.rows_expected:
            detail.append(f"Its row counts do not reconcile either: "
                          f"{meta.rows_expected:,} expected - "
                          f"{meta.duplicates_skipped:,} duplicates != "
                          f"{meta.rows_exported:,} exported.")
        detail.append("Re-run the extraction in the Amazon Ads console and let it "
                      "finish before downloading.")
        raise ValueError(" ".join(detail))

    qa.rows_parsed = len(events) + qa.rows_unparsable_time
    qa.distinct_campaigns = len(campaigns)
    qa.campaigns_with_budget_events = len(
        {e.campaign for e in events if e.machine == "budget"}
    )
    qa.date_keys = sorted(dates)
    return events, meta, qa


def event_identity(e: Event) -> tuple:
    """What makes a change row unique, independent of which export it came from.

    The entity name and the second matter. One campaign can pause seventeen
    different ad groups in the same minute -- those rows share everything except
    `level_name`, and merging them would silently delete real history.
    """
    return (e.date_key, e.second, e.level_type, e.level_name, e.campaign,
            e.change_type, e.from_val, e.to_val)


def dedupe_events(events: list[Event]) -> tuple[list[Event], int]:
    """Drop rows that appear in more than one export.

    Amazon's exports are date-range based, so loading a week and then a month
    that contains it is normal. Without this the overlap is not double-counted
    -- the state machine collapses the repeats -- but every repeated transition
    registers as a contradiction, burying the handful of genuine ones.
    """
    seen: set[tuple] = set()
    unique: list[Event] = []
    for e in events:
        identity = event_identity(e)
        if identity in seen:
            continue
        seen.add(identity)
        unique.append(e)
    return unique, len(events) - len(unique)


def discover_exports(*roots: str | Path) -> list[Path]:
    """Find `amazon-ads-history_*.xlsx` files, newest last, de-duplicated."""
    found: dict[Path, None] = {}
    for root in roots:
        root = Path(root)
        if root.is_file() and root.suffix.lower() in (".xlsx", ".xlsm"):
            found[root.resolve()] = None
        elif root.is_dir():
            for p in sorted(root.glob("*.xlsx")):
                if not p.name.startswith("~$"):
                    found[p.resolve()] = None
    return list(found)
